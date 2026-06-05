"""
Integration tests for adapter API workflows

Tests the API endpoints that support the adaptive import system,
including adapter discovery and role-based access control.
"""

import tempfile
from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook

from src.database.database_service import create_default_mocku_institution


@pytest.mark.integration
class TestAdapterAPIWorkflows:
    """Integration tests for adapter API endpoints."""

    def setup_method(self) -> None:
        """Set up test environment."""
        # Ensure MockU institution exists
        self.institution_id = create_default_mocku_institution()

    def test_site_admin_adapter_discovery_workflow(self, client: Any) -> None:
        """Test site admin can discover all available adapters via API."""
        # Mock site admin user session (compatible with new SessionService)
        with client.session_transaction() as sess:
            sess["user_id"] = "test-site-admin-123"
            sess["email"] = "admin@mocku.test"
            sess["role"] = "site_admin"
            sess["institution_id"] = self.institution_id
            sess["program_ids"] = []
            sess["display_name"] = "Test Site Admin"
            sess["created_at"] = "2024-01-01T00:00:00Z"

        # Test adapter discovery endpoint
        response = client.get("/api/adapters")
        assert response.status_code == 200

        data = response.get_json()
        assert data["success"] is True
        assert "adapters" in data
        assert len(data["adapters"]) >= 1

        # Should find CEI adapter (real customer adapter)
        cei_adapter = next(
            (a for a in data["adapters"] if a["id"] == "cei_excel_format_v1"), None
        )
        assert cei_adapter is not None
        assert cei_adapter["name"] == "CEI Excel Format v1.2"
        assert ".xlsx" in cei_adapter["supported_formats"]
        assert "courses" in cei_adapter["data_types"]

    def test_institution_admin_adapter_discovery_workflow(self, client: Any) -> None:
        """Test institution admin only sees their institution's adapters."""
        # Mock institution admin user session (compatible with new SessionService)
        with client.session_transaction() as sess:
            sess["user_id"] = "test-institution-admin-123"
            sess["email"] = "admin@mocku.test"
            sess["role"] = "institution_admin"
            sess["institution_id"] = self.institution_id
            sess["program_ids"] = []
            sess["display_name"] = "Test Institution Admin"
            sess["created_at"] = "2024-01-01T00:00:00Z"

        # Test adapter discovery endpoint
        response = client.get("/api/adapters")
        assert response.status_code == 200

        data = response.get_json()
        assert data["success"] is True
        assert "adapters" in data

        # Should only see adapters for their institution OR public adapters (institution_id=None)
        for adapter in data["adapters"]:
            assert adapter["institution_id"] in [
                self.institution_id,
                None,
            ], f"Inst admin should only see their institution's adapters or public adapters, got: {adapter['institution_id']}"

    def test_instructor_adapter_discovery_workflow(self, client: Any) -> None:
        """Test instructor sees no adapters (no import permissions)."""
        # Mock instructor user session
        with client.session_transaction() as sess:
            sess["user_id"] = "test-instructor-123"
            sess["email"] = "instructor@mocku.test"
            sess["role"] = "instructor"
            sess["institution_id"] = self.institution_id
            sess["program_ids"] = []
            sess["display_name"] = "Test Instructor"
            sess["created_at"] = "2024-01-01T00:00:00Z"

        # Test adapter discovery endpoint
        response = client.get("/api/adapters")
        assert response.status_code == 200

        data = response.get_json()
        assert data["success"] is True
        assert "adapters" in data

        # Instructors should see no adapters
        assert len(data["adapters"]) == 0

    def test_unauthenticated_adapter_access(self, client: Any) -> None:
        """Test that unauthenticated users cannot access adapter API."""
        # No user session - should be denied
        response = client.get("/api/adapters")
        assert response.status_code == 401

        data = response.get_json()
        assert data["success"] is False
        assert "authentication required" in data["error"].lower()

    def create_test_excel_file(self, file_path: str) -> None:
        """Create a minimal test Excel file."""
        workbook = Workbook()
        worksheet = workbook.active

        # MockU format headers
        headers = [
            "course",
            "section",
            "effterm_c",
            "students",
            "Faculty Name",
            "email",
        ]
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)

        # Sample data
        worksheet.cell(row=2, column=1, value="TEST-101")
        worksheet.cell(row=2, column=2, value="01")
        worksheet.cell(row=2, column=3, value="2024FA")
        worksheet.cell(row=2, column=4, value="10")
        worksheet.cell(row=2, column=5, value="Test Instructor")
        worksheet.cell(row=2, column=6, value="test@mocku.test")

        workbook.save(file_path)

    def test_site_admin_import_workflow_via_api(self, client: Any) -> None:
        """Test complete import workflow via API for site admin."""
        # Mock site admin user session
        with client.session_transaction() as sess:
            sess["user_id"] = "test-site_admin-123"
            sess["email"] = "admin@mocku.test"
            sess["role"] = "site_admin"
            sess["institution_id"] = self.institution_id
            sess["program_ids"] = []
            sess["display_name"] = "Test Site Admin"
            sess["created_at"] = "2024-01-01T00:00:00Z"

        with tempfile.TemporaryDirectory() as tmp_dir:
            # Create test file
            test_file = Path(tmp_dir) / "test_import.xlsx"
            self.create_test_excel_file(str(test_file))

            # Test import via API
            with open(test_file, "rb") as f:
                response = client.post(
                    "/api/import/excel",
                    data={
                        "excel_file": (f, "test_import.xlsx"),
                        "import_adapter": "cei_excel_format_v1",
                        "conflict_strategy": "use_theirs",
                        "dry_run": "false",
                    },
                )

            # Real imports run asynchronously: the request returns 202 + a
            # progress_id, and the result is reported via the progress tracker.
            assert response.status_code == 202
            data = response.get_json()
            assert data["success"] is True
            progress_id = data["progress_id"]

            import time

            progress = {}
            for _ in range(100):
                progress = client.get(
                    f"/api/import/progress/{progress_id}"
                ).get_json()
                if progress.get("status") in ("complete", "error"):
                    break
                time.sleep(0.05)

            assert progress.get("status") == "complete", progress
            assert progress["result"]["records_processed"] > 0

    def test_instructor_import_restriction_via_api(self, client: Any) -> None:
        """Test that instructors cannot import via API."""
        # Mock instructor user session
        with client.session_transaction() as sess:
            sess["user_id"] = "test-instructor-123"
            sess["email"] = "instructor@mocku.test"
            sess["role"] = "instructor"
            sess["institution_id"] = self.institution_id
            sess["program_ids"] = []
            sess["display_name"] = "Test Instructor"
            sess["created_at"] = "2024-01-01T00:00:00Z"

        with tempfile.TemporaryDirectory() as tmp_dir:
            # Create test file
            test_file = Path(tmp_dir) / "test_import.xlsx"
            self.create_test_excel_file(str(test_file))

            # Attempt import via API - should be denied
            with open(test_file, "rb") as f:
                response = client.post(
                    "/api/import/excel",
                    data={
                        "excel_file": (f, "test_import.xlsx"),
                        "import_adapter": "cei_excel_format_v1",
                        "conflict_strategy": "use_theirs",
                        "dry_run": "false",
                    },
                )

            # Should be denied based on role
            assert response.status_code in [403, 400]  # Forbidden or bad request
            data = response.get_json()
            assert data["success"] is False

    def test_adapter_metadata_consistency(self, client: Any) -> None:
        """Test that adapter metadata is consistent across API calls."""
        # Mock site admin user session
        with client.session_transaction() as sess:
            sess["user_id"] = "test-site_admin-123"
            sess["email"] = "admin@mocku.test"
            sess["role"] = "site_admin"
            sess["institution_id"] = self.institution_id
            sess["program_ids"] = []
            sess["display_name"] = "Test Site Admin"
            sess["created_at"] = "2024-01-01T00:00:00Z"

        # Get adapters via API
        response = client.get("/api/adapters")
        assert response.status_code == 200

        data = response.get_json()
        cei_adapter = next(
            (a for a in data["adapters"] if a["id"] == "cei_excel_format_v1"), None
        )
        assert cei_adapter is not None

        # Verify required metadata fields (note: institution_id may be None for global adapters like CEI)
        required_fields = [
            "id",
            "name",
            "description",
            "supported_formats",
            "data_types",
        ]
        for field in required_fields:
            assert field in cei_adapter
            assert cei_adapter[field] is not None

        # Verify data types are sensible
        assert isinstance(cei_adapter["data_types"], list)
        assert len(cei_adapter["data_types"]) > 0

        # Verify supported formats
        assert isinstance(cei_adapter["supported_formats"], list)
        assert ".xlsx" in cei_adapter["supported_formats"]

    def test_error_handling_in_api_workflows(self, client: Any) -> None:
        """Test API error handling for various failure scenarios."""
        # Mock site admin user session
        with client.session_transaction() as sess:
            sess["user_id"] = "test-site_admin-123"
            sess["email"] = "admin@mocku.test"
            sess["role"] = "site_admin"
            sess["institution_id"] = self.institution_id
            sess["program_ids"] = []
            sess["display_name"] = "Test Site Admin"
            sess["created_at"] = "2024-01-01T00:00:00Z"

        with tempfile.TemporaryDirectory() as tmp_dir:
            # Test 1: Invalid file format
            invalid_file = Path(tmp_dir) / "invalid.txt"
            invalid_file.write_text("This is not an Excel file")

            with open(invalid_file, "rb") as f:
                response = client.post(
                    "/api/import/excel",
                    data={
                        "excel_file": (f, "invalid.txt"),
                        "import_adapter": "cei_excel_format_v1",
                        "conflict_strategy": "use_theirs",
                        "dry_run": "false",
                    },
                )

            assert response.status_code == 400  # Bad request
            data = response.get_json()
            assert data["success"] is False
            assert "error" in data

            # Test 2: Non-existent adapter
            test_file = Path(tmp_dir) / "test.xlsx"
            self.create_test_excel_file(str(test_file))

            with open(test_file, "rb") as f:
                response = client.post(
                    "/api/import/excel",
                    data={
                        "excel_file": (f, "test.xlsx"),
                        "import_adapter": "nonexistent_adapter",
                        "conflict_strategy": "use_theirs",
                        "dry_run": "false",
                    },
                )

            assert response.status_code == 400  # Bad request
            data = response.get_json()
            assert data["success"] is False
            assert (
                "adapter not found" in data["error"].lower()
                or "access denied" in data["error"].lower()
                or "target institution" in data["error"].lower()
            )

    def test_dry_run_workflow_via_api(self, client: Any) -> None:
        """Test dry run functionality via API."""
        # Mock site admin user session
        with client.session_transaction() as sess:
            sess["user_id"] = "test-site_admin-123"
            sess["email"] = "admin@mocku.test"
            sess["role"] = "site_admin"
            sess["institution_id"] = self.institution_id
            sess["program_ids"] = []
            sess["display_name"] = "Test Site Admin"
            sess["created_at"] = "2024-01-01T00:00:00Z"

        with tempfile.TemporaryDirectory() as tmp_dir:
            # Create test file
            test_file = Path(tmp_dir) / "test_dry_run.xlsx"
            self.create_test_excel_file(str(test_file))

            # Test dry run import
            with open(test_file, "rb") as f:
                response = client.post(
                    "/api/import/excel",
                    data={
                        "excel_file": (f, "test_dry_run.xlsx"),
                        "import_adapter": "cei_excel_format_v1",
                        "conflict_strategy": "use_theirs",
                        "dry_run": "true",
                    },
                )

            assert response.status_code == 200
            data = response.get_json()
            assert data["success"] is True
            assert "dry_run" in data or "preview" in data.get("message", "").lower()

    def test_cross_role_adapter_access_patterns(self, client: Any) -> None:
        """Test adapter access patterns across different user roles."""
        roles_and_expected_access = [
            ("site_admin", True, "Should see all adapters"),
            ("institution_admin", True, "Should see institution adapters"),
            ("program_admin", True, "Should see institution adapters"),
            ("instructor", False, "Should see no adapters"),
        ]

        for role, should_have_access, description in roles_and_expected_access:
            with client.session_transaction() as sess:
                sess["user_id"] = f"test-{role}-123"
                sess["email"] = f"{role}@mocku.test"
                sess["role"] = role
                sess["institution_id"] = self.institution_id
                sess["program_ids"] = []
                sess["display_name"] = f"Test {role.replace('_', ' ').title()}"
                sess["created_at"] = "2024-01-01T00:00:00Z"

            response = client.get("/api/adapters")
            assert response.status_code == 200

            data = response.get_json()
            assert data["success"] is True

            if should_have_access:
                assert (
                    len(data["adapters"]) > 0
                ), f"{description}: {role} should have adapter access"
            else:
                assert (
                    len(data["adapters"]) == 0
                ), f"{description}: {role} should have no adapter access"
